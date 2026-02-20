"""
═══════════════════════════════════════════════════════════════
  generate_tasks_gradebook.py — مولّد جدول درجات المهام
  يقرأ simulation_data.json ويولّد درجات 5 مهام (M1→M5) لكل طالب
  مع منحنى تعلّم واقعي، ديناميكيات مجموعات، وعقوبات التأخير.

  الناتج: tasks_gradebook.xlsx

  الاستخدام:
    python generate_tasks_gradebook.py
    python generate_tasks_gradebook.py --input simulation_data.json
    python generate_tasks_gradebook.py --input sim.json --output grades.xlsx --seed 42
═══════════════════════════════════════════════════════════════
"""

import sys
import io

# ضمان عرض النصوص العربية بشكل صحيح على Windows
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "buffer"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import json
import argparse
import os
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ════════════════════════════════════════════════════════════════
#  الإعدادات الثابتة
# ════════════════════════════════════════════════════════════════

TASK_CONFIG = {
    # أسماء المهام
    "tasks": ["M1", "M2", "M3", "M4", "M5"],

    # مواعيد التسليم النهائية (بعد 2 أسبوع من بداية الكورس)
    "deadlines": [
        "2026-01-24",   # M1 — الأسبوع 2
        "2026-02-07",   # M2 — الأسبوع 4
        "2026-02-21",   # M3 — الأسبوع 6
        "2026-03-07",   # M4 — الأسبوع 8
        "2026-03-21",   # M5 — الأسبوع 10
    ],

    # نافذة التسليم: يسلّم الطالب في الـ X يوم قبل الموعد
    "submitWindowDays": 5,

    # المجموعات التي تخضع لعقوبة التأخير (محدّدة الوقت)
    "lateGroups": ["G2", "G4"],
    # عقوبة التأخير حسب الإطار النظري:
    # 0-6 ساعات: خصم 30% | 6-24 ساعة: خصم 60% | >24 ساعة: صفر

    # المجموعات التشاركية
    "collaborativeGroups": ["G3", "G4"],
    "collaborativeBonus": 3,    # +3 نقطة مكافأة
    "collaborativeNoiseSD": 4,  # تشويش أقل
    "collaborativeLevelingFactor": 0.15,  # 15% نحو متوسط المجموعة

    # المجموعات التنافسية
    "competitiveGroups": ["G1", "G2"],
    "competitiveNoiseSD": 8,    # تشويش أعلى

    # تأثير الانخراط (Flow) على الدرجة
    "flowWeight": 6,            # (postFlowLevel - 0.5) × 6 → نطاق -3 إلى +3

    # نطاق الدرجات الأساسية بعد تحويل المهارة
    "scoreMin": 50,             # أدنى درجة لمهارة = 0
    "scoreMax": 95,             # أعلى درجة لمهارة = 1

    # الدرجة القصوى لكل مهمة
    "taskMaxScore": 100,
}

# الطلاب المتسربون — يكملون M1 و M2 فقط
DROPOUT_IDS = [
    "STD-081", "STD-082", "STD-083", "STD-084",
    "STD-085", "STD-086", "STD-087", "STD-088",
    "STD-089", "STD-090", "STD-091", "STD-092",
    "STD-093", "STD-094", "STD-095", "STD-096",
]

# تسميات الحروف
GRADE_THRESHOLDS = [
    (90, "A+"), (85, "A"), (80, "B+"), (75, "B"),
    (70, "C+"), (65, "C"), (60, "D+"), (55, "D"), (0, "F"),
]


# ════════════════════════════════════════════════════════════════
#  قراءة البيانات
# ════════════════════════════════════════════════════════════════

def load_data(json_path):
    """يقرأ simulation_data.json ويُعيد قائمة الطلاب."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    students = data["students"]
    print(f"[OK] تم تحميل {len(students)} طالب من {os.path.basename(json_path)}")
    return students


def inject_dropouts(students):
    """يضيف الطلاب المتسربين (16) إن لم يكونوا في الـ JSON ليصبح العدد 96."""
    existing_ids = {s["id"] for s in students}
    added = 0
    for d_id in DROPOUT_IDS:
        if d_id not in existing_ids:
            num = int(d_id.split("-")[1])
            if num <= 84:
                grp = "G1"
            elif num <= 88:
                grp = "G2"
            elif num <= 92:
                grp = "G3"
            else:
                grp = "G4"
            students.append({
                "id": d_id, "name": f"طالب {d_id}",
                "email": f"{d_id.lower().replace('-', '')}@student.edu",
                "group": grp,
                "preSkill": 0.25, "postSkill": 0.30,
                "postFlowLevel": 0.20,
            })
            added += 1
    if added > 0:
        print(f"[Dropouts] تمت اضافة {added} طالب متسرب (الاجمالي: {len(students)})")


def assign_teams(students):
    """يقسّم طلاب G3/G4 إلى فرق (كل 4 طلاب متتاليين = فريق)."""
    team_map = {}
    counters = {"G3": 0, "G4": 0}
    for s in students:
        grp = s["group"]
        if grp in ("G3", "G4"):
            team_idx = counters[grp] // 4 + 1
            team_map[s["id"]] = f"فريق {team_idx}"
            counters[grp] += 1
        else:
            team_map[s["id"]] = "عمل فردي"
    return team_map


# ════════════════════════════════════════════════════════════════
#  دوال توليد الدرجات
# ════════════════════════════════════════════════════════════════

def skill_to_base(skill):
    """يُحوّل مستوى المهارة [0–1] إلى درجة أساسية [scoreMin–scoreMax]."""
    lo = TASK_CONFIG["scoreMin"]
    hi = TASK_CONFIG["scoreMax"]
    return lo + skill * (hi - lo)


def get_task_noise_sd(group):
    """يُعيد الانحراف المعياري للتشويش بحسب نوع المجموعة."""
    if group in TASK_CONFIG["collaborativeGroups"]:
        return TASK_CONFIG["collaborativeNoiseSD"]
    return TASK_CONFIG["competitiveNoiseSD"]


def get_collaborative_bonus(group):
    """يُعيد المكافأة الإضافية للمجموعات التشاركية."""
    if group in TASK_CONFIG["collaborativeGroups"]:
        return TASK_CONFIG["collaborativeBonus"]
    return 0.0


def is_dropout(student_id):
    """يتحقق إذا كان الطالب متسرباً."""
    return student_id in DROPOUT_IDS


def generate_student_tasks(student, rng):
    """
    يولّد نتائج المهام لطالب واحد (أو متوسط فريق).
    التأخير مُستنتَج من postFlowLevel (تدفق عالٍ = التزام، منخفض = تأخير).
    """
    pre_skill  = student["preSkill"] or 0.25
    post_skill = student["postSkill"] or (pre_skill * 1.05)  # fallback للمتسربين
    post_flow  = student.get("postFlowLevel") or 0.30        # fallback للمتسربين
    group      = student["group"]
    sid        = student["id"]

    noise_sd     = get_task_noise_sd(group)
    collab_bonus = get_collaborative_bonus(group)
    flow_adj     = (post_flow - 0.5) * TASK_CONFIG["flowWeight"]
    dropout      = is_dropout(sid)

    results   = []
    num_tasks = len(TASK_CONFIG["tasks"])

    for task_idx in range(num_tasks):
        task_name = TASK_CONFIG["tasks"][task_idx]

        if dropout and task_idx >= 2:
            results.append({
                "task": task_name, "score": None, "raw_score": None,
                "is_late": None, "hours_late": 0, "submit_date": None,
            })
            continue

        task_progress = task_idx / (num_tasks - 1)
        skill_i = pre_skill + task_progress * (post_skill - pre_skill)

        base  = skill_to_base(skill_i)
        noise = rng.normal(0, noise_sd)
        raw   = base + collab_bonus + flow_adj + noise
        raw   = float(np.clip(raw, 0, TASK_CONFIG["taskMaxScore"]))

        # منطق تحديد التأخير لمجموعات الضغط الزمني (G2, G4)
        hours_late = 0
        late_flag = False

        if group in TASK_CONFIG["lateGroups"]:
            # التدفق العالي = تأخير أقل، التدفق المنخفض = تأخير أكبر
            # المتوسط السالب لأغلب الطلاب = تسليم في الوقت
            expected_delay_hours = -10.0 + (1.0 - post_flow) * 20.0
            delay = rng.normal(expected_delay_hours, 3.5)

            if delay > 0:
                hours_late = round(delay, 1)
                late_flag = True

        results.append({
            "task": task_name, "score": raw, "raw_score": raw,
            "is_late": late_flag, "hours_late": hours_late, "submit_date": None,
        })

    return results


# ════════════════════════════════════════════════════════════════
#  بناء النتائج (فرق تشاركية + أفراد + توقيتات مرتبطة بالتدفق)
# ════════════════════════════════════════════════════════════════

def generate_all_results(students, team_map, rng):
    """
    يولّد نتائج المهام لجميع الطلاب:
    - G3/G4 التشاركية: نتيجة واحدة لكل فريق تُنسخ لجميع أعضائه
    - G1/G2 التنافسية: نتيجة فردية لكل طالب
    التوقيتات مرتبطة بـ postFlowLevel (تدفق عالٍ = تسليم مبكر).
    """
    all_results = [None] * len(students)
    team_cache = {}
    deadlines = [datetime.strptime(d, "%Y-%m-%d") for d in TASK_CONFIG["deadlines"]]

    for i, student in enumerate(students):
        grp  = student["group"]
        sid  = student["id"]
        team = team_map[sid]
        drop = is_dropout(sid)

        if team != "عمل فردي":
            team_key = f"{grp}_{team}"
            if team_key not in team_cache:
                members  = [s for s in students
                            if team_map[s["id"]] == team and s["group"] == grp]
                avg_pre  = float(np.mean([m["preSkill"] or 0.25 for m in members]))
                avg_post = float(np.mean([(m["postSkill"] or (m["preSkill"] or 0.25) * 1.05) for m in members]))
                avg_flow = float(np.mean([(m.get("postFlowLevel") or 0.30) for m in members]))
                team_student = {"id": "TEAM", "group": grp,
                                "preSkill": avg_pre, "postSkill": avg_post,
                                "postFlowLevel": avg_flow}
                team_res = generate_student_tasks(team_student, rng)
                _assign_timestamps_for_student(team_res, grp, avg_flow, deadlines, rng)
                team_cache[team_key] = team_res

            base = team_cache[team_key]
            results = []
            for t_idx, res in enumerate(base):
                if drop and t_idx >= 2:
                    results.append({"task": res["task"], "score": None, "raw_score": None,
                                    "is_late": None, "hours_late": 0, "submit_date": None})
                else:
                    results.append(dict(res))
            all_results[i] = results
        else:
            res = generate_student_tasks(student, rng)
            _assign_timestamps_for_student(
                res, grp, student.get("postFlowLevel") or 0.30, deadlines, rng)
            all_results[i] = res

    return all_results


def _assign_timestamps_for_student(results, group, post_flow, deadlines, rng):
    """يُعيّن توقيتات تسليم مرتبطة بالتدفق: تدفق عالٍ = تسليم مبكر."""
    window  = TASK_CONFIG["submitWindowDays"]
    n_tasks = len(TASK_CONFIG["tasks"])

    for task_idx, res in enumerate(results):
        if res["score"] is None:
            continue

        deadline = deadlines[task_idx]

        if group in TASK_CONFIG["lateGroups"]:
            if res["is_late"] and res.get("hours_late", 0) > 0:
                # إضافة التأخير بالساعات فوق الموعد النهائي
                submit_time = deadline + timedelta(hours=res["hours_late"])
                # تجنب تسليمات الفجر (بين 2 صباحاً و 8 صباحاً)
                if 2 <= submit_time.hour <= 8:
                    submit_time -= timedelta(hours=7)
            else:
                # تسليم مبكر
                early_hours = (post_flow * 48) + rng.normal(0, 10)
                early_hours = max(1, min(window * 24, early_hours))
                submit_time = deadline - timedelta(hours=early_hours)

            res["submit_date"] = submit_time.strftime("%Y-%m-%d %H:%M")
        else:
            if task_idx == n_tasks - 1:
                submit_day = deadline - timedelta(days=float(rng.uniform(0, 1.5)))
            else:
                procrastination = (1.0 - post_flow) * window
                base_early = window - procrastination + rng.normal(0, 1.0)
                base_early = max(0, min(window, base_early))
                submit_day = deadline - timedelta(days=base_early)

            hour   = int(rng.integers(9, 23))
            minute = int(rng.integers(0, 60))
            submit_dt = submit_day.replace(hour=hour, minute=minute, second=0)
            res["submit_date"] = submit_dt.strftime("%Y-%m-%d %H:%M")


def apply_late_penalties(all_results):
    """
    يطبّق عقوبة التأخير بناءً على الإطار النظري:
    - تأخير (0 إلى 6 ساعات): خصم 30%
    - تأخير (6 إلى 24 ساعة): خصم 60%
    - تأخير (أكثر من 24 ساعة): المهمة 0
    """
    for student_results in all_results:
        for res in student_results:
            if res["score"] is None or not res["is_late"]:
                continue

            # التأخير بالساعات
            delay_hours = res.get("hours_late", 0)

            if 0 < delay_hours <= 6:
                # خصم 30%
                res["score"] = float(np.clip(res["score"] * 0.70, 0, TASK_CONFIG["taskMaxScore"]))
            elif 6 < delay_hours <= 24:
                # خصم 60%
                res["score"] = float(np.clip(res["score"] * 0.40, 0, TASK_CONFIG["taskMaxScore"]))
            elif delay_hours > 24:
                # صفر
                res["score"] = 0.0


# ════════════════════════════════════════════════════════════════
#  بناء DataFrame
# ════════════════════════════════════════════════════════════════

def letter_grade(pct):
    """يُحوّل النسبة المئوية إلى حرف تقديري."""
    if pct is None:
        return "—"
    for threshold, letter in GRADE_THRESHOLDS:
        if pct >= threshold:
            return letter
    return "F"


def build_dataframe(all_results, students, team_map):
    """يبني DataFrame من نتائج المهام مع عمود الفريق والبونص."""
    rows = []
    tasks = TASK_CONFIG["tasks"]

    for i, student in enumerate(students):
        sid = student["id"]
        row = {
            "ID":    sid,
            "Name":  student["name"],
            "Group": student["group"],
            "Team":  team_map.get(sid, "عمل فردي"),
        }

        task_scores = []
        for task_idx, task_name in enumerate(tasks):
            res = all_results[i][task_idx]
            score = round(res["score"], 1) if res["score"] is not None else None
            row[task_name]                = score
            row[f"{task_name}_Date"]      = res["submit_date"] if score is not None else None
            row[f"{task_name}_Late"]      = ("نعم" if res["is_late"] else "لا") if score is not None else None
            if score is not None:
                task_scores.append(score)

        max_possible = len(task_scores) * TASK_CONFIG["taskMaxScore"]
        total  = round(sum(task_scores), 1) if task_scores else 0
        pct    = round((total / max_possible * 100), 1) if max_possible > 0 else None

        row["Bonus"]       = 0
        row["Total"]       = total
        row["Max_Possible"] = max_possible
        row["Percentage"]  = pct
        row["Grade"]       = letter_grade(pct)
        row["Is_Dropout"]  = "نعم" if is_dropout(sid) else "لا"

        rows.append(row)

    return pd.DataFrame(rows)


def apply_competitive_bonuses(df):
    """يطبّق مكافآت ترتيبية لـ G1 (حسب الدرجة) و G2 (حسب السرعة + الجودة)."""
    tasks = TASK_CONFIG["tasks"]

    g1_active = df[(df["Group"] == "G1") & (df["Is_Dropout"] == "لا")]
    g1_sorted = g1_active.sort_values("Total", ascending=False).index
    for rank, idx in enumerate(g1_sorted):
        if rank < 3:
            df.at[idx, "Bonus"] = 15
        elif rank < 8:
            df.at[idx, "Bonus"] = 5

    g2_active = df[(df["Group"] == "G2") & (df["Is_Dropout"] == "لا")]
    g2_data = []
    for idx in g2_active.index:
        late_count = sum(1 for t in tasks if df.at[idx, f"{t}_Late"] == "نعم")
        g2_data.append((idx, late_count, df.at[idx, "Total"]))
    g2_data.sort(key=lambda x: (x[1], -x[2]))
    for rank, (idx, _, total) in enumerate(g2_data):
        if rank < 3 and total >= 300:
            df.at[idx, "Bonus"] = 20
        elif rank < 6:
            df.at[idx, "Bonus"] = 15
        elif rank < 10:
            df.at[idx, "Bonus"] = 10

    mask = df["Bonus"] > 0
    df.loc[mask, "Total"] = df.loc[mask, "Total"] + df.loc[mask, "Bonus"]
    df.loc[mask, "Percentage"] = (
        df.loc[mask, "Total"] / df.loc[mask, "Max_Possible"] * 100
    ).round(1)
    df.loc[mask, "Grade"] = df.loc[mask, "Percentage"].apply(letter_grade)


# ════════════════════════════════════════════════════════════════
#  حفظ Excel مع التنسيق
# ════════════════════════════════════════════════════════════════

# ألوان التنسيق
FILL_HEADER     = PatternFill("solid", fgColor="2E4057")  # كحلي غامق
FILL_LATE       = PatternFill("solid", fgColor="FFD6A5")  # برتقالي فاتح
FILL_DROPOUT    = PatternFill("solid", fgColor="E8E8E8")  # رمادي فاتح
FILL_EMPTY      = PatternFill("solid", fgColor="F0F0F0")  # رمادي للخلايا الفارغة
FILL_GRADE_A    = PatternFill("solid", fgColor="B7E4C7")  # أخضر للـ A
FILL_GRADE_F    = PatternFill("solid", fgColor="FFCCD5")  # أحمر للـ F
FILL_GROUP      = {
    "G1": PatternFill("solid", fgColor="D0E8FF"),  # أزرق فاتح
    "G2": PatternFill("solid", fgColor="FFE4E1"),  # وردي فاتح
    "G3": PatternFill("solid", fgColor="DFFFD8"),  # أخضر فاتح
    "G4": PatternFill("solid", fgColor="FFF9C4"),  # أصفر فاتح
}
FONT_HEADER     = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_DROPOUT    = Font(name="Calibri", italic=True, color="888888")

THIN_BORDER_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER      = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE,
)


def save_excel(df, output_path):
    """يحفظ DataFrame كـ Excel مع تنسيق احترافي."""
    # حفظ أولي بـ pandas
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Gradebook")

    # فتح الملف للتنسيق المتقدم
    wb = load_workbook(output_path)
    ws = wb["Gradebook"]

    tasks     = TASK_CONFIG["tasks"]
    num_cols  = ws.max_column
    num_rows  = ws.max_row

    # ─── تنسيق رأس الجدول ───────────────────────────────────────
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # ─── بناء خريطة العمود ─────────────────────────────────────
    col_map = {ws.cell(row=1, column=c).value: c for c in range(1, num_cols + 1)}

    # ─── تنسيق صفوف البيانات ───────────────────────────────────
    for row_idx in range(2, num_rows + 1):
        group     = ws.cell(row=row_idx, column=col_map.get("Group", 3)).value
        is_drop   = ws.cell(row=row_idx, column=col_map.get("Is_Dropout", num_cols)).value == "نعم"
        grade_val = ws.cell(row=row_idx, column=col_map.get("Grade", num_cols)).value

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # خلايا فارغة (متسربون M3-M5)
            if cell.value is None:
                cell.fill = FILL_EMPTY
                continue

            # لون صف المتسرب
            if is_drop:
                cell.font = FONT_DROPOUT

            # لون حسب المجموعة (عمودا ID و Group فقط)
            header_name = ws.cell(row=1, column=col_idx).value
            if header_name in ("ID", "Name", "Group", "Team") and group in FILL_GROUP:
                cell.fill = FILL_GROUP[group]

        # تلوين خلايا التأخير
        for task_name in tasks:
            late_col = col_map.get(f"{task_name}_Late")
            score_col = col_map.get(task_name)
            date_col  = col_map.get(f"{task_name}_Date")
            if late_col and ws.cell(row=row_idx, column=late_col).value == "نعم":
                for c in [score_col, date_col, late_col]:
                    if c:
                        ws.cell(row=row_idx, column=c).fill = FILL_LATE

        # تلوين خلية الدرجة
        grade_col = col_map.get("Grade")
        if grade_col and grade_val:
            if isinstance(grade_val, str) and grade_val.startswith("A"):
                ws.cell(row=row_idx, column=grade_col).fill = FILL_GRADE_A
            elif grade_val == "F":
                ws.cell(row=row_idx, column=grade_col).fill = FILL_GRADE_F

    # ─── عرض الأعمدة ───────────────────────────────────────────
    col_widths = {
        "ID": 12, "Name": 22, "Group": 8, "Team": 14,
        "Bonus": 8, "Total": 10, "Max_Possible": 12, "Percentage": 10,
        "Grade": 8, "Is_Dropout": 10,
    }
    for task_name in tasks:
        col_widths[task_name]             = 8
        col_widths[f"{task_name}_Date"]   = 18
        col_widths[f"{task_name}_Late"]   = 8

    for col_idx in range(1, num_cols + 1):
        header = ws.cell(row=1, column=col_idx).value
        width  = col_widths.get(header, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ─── تجميد الصف الأول والأعمدة الثلاثة ─────────────────────
    ws.freeze_panes = "E2"

    # ─── ورقة ملخص المجموعات ────────────────────────────────────
    ws_summary = wb.create_sheet("Group_Summary")
    _write_group_summary(ws_summary, df)

    wb.save(output_path)
    print(f"[Excel] تم الحفظ في: {output_path}")


def _write_group_summary(ws, df):
    """يكتب ورقة ملخص إحصائي لكل مجموعة."""
    tasks = TASK_CONFIG["tasks"]
    headers = ["Group", "Type", "N", "N_Active"] + \
              [f"{t}_Mean" for t in tasks] + \
              ["Total_Mean", "Pct_Mean", "Late_Count"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    group_types = {"G1": "تنافسي+مفتوح", "G2": "تنافسي+محدد",
                   "G3": "تشاركي+مفتوح", "G4": "تشاركي+محدد"}

    for row_idx, grp in enumerate(["G1", "G2", "G3", "G4"], 2):
        gdf        = df[df["Group"] == grp]
        active_df  = gdf[gdf["Is_Dropout"] == "لا"]
        late_count = sum(
            1 for t in tasks
            for _, r in gdf.iterrows()
            if r.get(f"{t}_Late") == "نعم"
        )
        row_data = [
            grp,
            group_types.get(grp, ""),
            len(gdf),
            len(active_df),
        ]
        for t in tasks:
            vals = pd.to_numeric(gdf[t], errors="coerce").dropna()
            row_data.append(round(vals.mean(), 2) if len(vals) > 0 else None)
        row_data.append(round(pd.to_numeric(gdf["Total"], errors="coerce").mean(), 2))
        row_data.append(round(pd.to_numeric(gdf["Percentage"], errors="coerce").mean(), 2))
        row_data.append(late_count)

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.alignment = Alignment(horizontal="center")
            if grp in FILL_GROUP:
                cell.fill = FILL_GROUP[grp]

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


# ════════════════════════════════════════════════════════════════
#  الدالة الرئيسية
# ════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="يولّد جدول درجات المهام من simulation_data.json")
    parser.add_argument("--input",  default="simulation_data.json",
                        help="مسار ملف simulation_data.json (افتراضي: simulation_data.json)")
    parser.add_argument("--output", default="tasks_gradebook.xlsx",
                        help="مسار ملف Excel الناتج (افتراضي: tasks_gradebook.xlsx)")
    parser.add_argument("--seed",   type=int, default=None,
                        help="Seed للعشوائية (لإعادة الإنتاج)")
    args = parser.parse_args()

    # ─── تحديد المسارات ─────────────────────────────────────────
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    input_path  = args.input if os.path.isabs(args.input) else os.path.join(script_dir, args.input)
    output_path = args.output if os.path.isabs(args.output) else os.path.join(script_dir, args.output)

    if not os.path.exists(input_path):
        print(f"❌ الملف غير موجود: {input_path}")
        return

    # ─── تهيئة العشوائية ────────────────────────────────────────
    seed = args.seed if args.seed is not None else 42
    rng  = np.random.default_rng(seed)
    print(f"[Seed] {seed}")

    # ─── تحميل البيانات ─────────────────────────────────────────
    students = load_data(input_path)

    # ─── حقن المتسربين إن لم يكونوا في الـ JSON ────────────────
    inject_dropouts(students)

    # ─── تعيين الفرق (G3/G4 → فرق، G1/G2 → عمل فردي) ─────────
    team_map = assign_teams(students)
    team_count = sum(1 for v in team_map.values() if v != "عمل فردي")
    print(f"[Teams] {team_count} طالب في فرق تشاركية")

    # ─── توليد النتائج (فرق + أفراد + توقيتات مرتبطة بالتدفق) ──
    print("[Generate] جاري توليد درجات المهام والتوقيتات...")
    all_results = generate_all_results(students, team_map, rng)

    # ─── تطبيق عقوبات التأخير ───────────────────────────────────
    print("[Penalty] تطبيق عقوبات التاخير لـ G2/G4...")
    apply_late_penalties(all_results)

    # ─── بناء DataFrame ─────────────────────────────────────────
    print("[DataFrame] بناء جدول البيانات...")
    df = build_dataframe(all_results, students, team_map)

    # ─── مكافآت تنافسية لـ G1/G2 ────────────────────────────────
    print("[Bonus] تطبيق المكافآت التنافسية لـ G1/G2...")
    apply_competitive_bonuses(df)

    # ─── ملخص إحصائي ────────────────────────────────────────────
    late_count = sum(1 for res_list in all_results
                     for res in res_list if res.get("is_late"))
    print("\n" + "=" * 44)
    print("  ملخص جدول الدرجات")
    print("=" * 44)
    print(f"  اجمالي الطلاب  : {len(df)}")
    print(f"  طلاب فاعلون    : {len(df[df['Is_Dropout'] == 'لا'])}")
    print(f"  متسربون         : {len(df[df['Is_Dropout'] == 'نعم'])}")
    print(f"  متاخرون          : {late_count}")
    print()
    for grp in ["G1", "G2", "G3", "G4"]:
        gdf = df[df["Group"] == grp]
        pcts = pd.to_numeric(gdf["Percentage"], errors="coerce").dropna()
        if len(pcts) > 0:
            print(f"  {grp}: متوسط = {pcts.mean():.1f}% | انحراف = {pcts.std():.1f}%")
    print("=" * 44)

    # ─── حفظ Excel ──────────────────────────────────────────────
    save_excel(df, output_path)
    print(f"\n[Done] الملف محفوظ في: {output_path}")


if __name__ == "__main__":
    main()
